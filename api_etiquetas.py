from flask import Flask
import psycopg2
import sys
import paramiko
import re
import uuid
from datetime import datetime
from flask_cors import CORS, cross_origin
import datetime
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import email_to
import sys
from json import loads

app = Flask(__name__)
CORS(app, support_credentials=True)

""" 
#    API´S Framework Flask
#    Franco Cumplido 
#    ECO S.A
#    2021


         ↓↓↓↓↓↓↓    Las siguientes API´S están vinculadas al Frontend del sistema de inventario (Edición vista formulario) Requisito → Tener activada "opciones de desarrollador":   ↓↓↓↓↓↓↓

impresion_out           = Se imrpimen etiquetas de ordenes de despacho.
impresion_picking       = Se imprimen etiquetas referentes a las ordenes de entrega("picking") 
impresion_picking_final = Se imprimen etiquetas del proceso PICKING-FINAL(fase del proceso donde se crean bultos).
delete_packages         = Si el trabajador o embalador se equivoca en la fase de embalaje y este anula las reservas se eliminan los paquetes y se resetea su contabilización.
insert_packages         = Al momento de validar la operación "PICKING-FINAL" con todos los bultos creados, estos se referencian en el Sistema Imatronix y la nota de venta queda como "Bultos creados" 
delete_specific_package** = Al eliminr un paquete en especifico del sistemoa 154 este debe eliminarse tambien del 200.
insert_codigos_de_barra = Este bloque de codigo ingresará los códigos de barras de las cajas master de productos TRUPER y derivados.
"""


@app.route('/impresion_out/<codigo_nota_de_venta>', methods=['GET', 'POST'])
def impresion_out(codigo_nota_de_venta):
    # funcion para reemplazar valores string
    def multiple_replace(string, rep_dict):
        pattern = re.compile("|".join([re.escape(k) for k in sorted(rep_dict, key=len, reverse=True)]), flags=re.DOTALL)
        return pattern.sub(lambda x: rep_dict[x.group(0)], string)

    # codigo_nota_de_venta = sys.argv[1]
    codigo_nota_de_venta = codigo_nota_de_venta
    #codigo_nota_de_venta = codigo_nota_de_venta[19:]
    # fecha_despacho = datetime.now()
    fecha_despacho = datetime.datetime.now()
    nombre_archivo = str(uuid.uuid4()) + ".zpl"

    conexion = psycopg2.connect("host=192.168.0.200 dbname=imatronix_ecosa user=postgres password=  ")
    cursor = conexion.cursor()
    query = f''' SELECT ''' \
            ''' cliente.razon_social, ''' \
            ''' ciudad.ciudad, ''' \
            ''' comuna.comuna, ''' \
            ''' region.region, ''' \
            ''' nota_de_venta.transporte_2, ''' \
            ''' nota_de_venta.direccion ''' \
            ''' FROM nota_de_venta ''' \
            ''' JOIN cliente ON nota_de_venta.cliente = cliente._id ''' \
            ''' JOIN ciudad  ON nota_de_venta.ciudad  = ciudad._id ''' \
            ''' JOIN comuna  ON nota_de_venta.comuna  = comuna._id ''' \
            ''' JOIN region  ON nota_de_venta.region  = region._id ''' \
            ''' WHERE nota_de_venta.codigo = %s; '''
    cursor.execute(query, (codigo_nota_de_venta,))
    datos = cursor.fetchall()

    for razon_social, ciudad, comuna, region, transporte_2, direccion in datos:

        if len(direccion) > 37:
            direccion_1 = direccion[:30]
            direccion_2 = direccion[30:]

        else:
            direccion_1 = direccion
            direccion_2 = " "

        valores_reemplazo = {'Ñ': 'N', 'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U', 'º': '#', '°': '#'}
        razon_social = multiple_replace(razon_social, valores_reemplazo)
        ciudad = multiple_replace(ciudad, valores_reemplazo)
        comuna = multiple_replace(comuna, valores_reemplazo)
        region = multiple_replace(region, valores_reemplazo)
        transporte_2 = multiple_replace(transporte_2, valores_reemplazo)
        direccion_1 = multiple_replace(direccion_1, valores_reemplazo)
        direccion_2 = multiple_replace(direccion_2, valores_reemplazo)

        argumento = "BEC1/ORDEN-ENTREGA/["+codigo_nota_de_venta+"]"
        conexion_154 = psycopg2.connect("host=192.168.0.154 dbname=InventorySystem user=postgres password=  ")
        cursor_154 = conexion_154.cursor()
        consulta = f''' SELECT note FROM stock_picking WHERE name = %s;'''

        cursor_154.execute(consulta, (argumento,))
        datos = cursor_154.fetchall()
        nota = str(datos[0][0])
        ini = (nota.find('Cantidad bultos') + 16)
        fin = nota.find('\n')
        cantidad_bultos = str(nota[ini:fin]).replace(' ','')


        # BLOQUE DE IMPRESION DEL DOUMENTO!  //ESTO DE DEBE CAMBIAR POR EL SERVIDOR DE IMPRESIÓN FINAL!
        host = "192.168.0.205"
        port = 22
        username = "root"
        password = "mcpt14a12"

        command = "^XA" \
                  "^FX SECCION DATOS EMPRESA." \
                  "^CF0,60" \
                  "^FO50,30^GFA,5125,5125,25,,::::::::g07FC,M0UFCI0UFE,L07UFCI0VFC,K01VFCI0WF,K07VFCI0WFC,J01WFCI0WFE,J03WFCI0XF,J07WFCI0XFC," \
                  "J0XFCI0XFC,I01XFCI0XFE,I03XFCI0YF,I07XFCI0YF8,I07XFCI0YFC,I0YFCI0YFC,I0YFCI0YFE,001 YFCI0YFE,001 YFCI0gF,003 YFCI0gF,:::007 YFCI0gF8," \
                  "::::::::::007 YFC001gF8,007 YFCI0gF8,: 007 YF8I07YF8,007 XF8K07XF8,007 WF8M0XF8,007 VFEN01WF8,007 VFP03VF8,007UFCQ0VF8,007UFR03UF8,007" \
                  "TFES0UF8,007 TF8K03FEK07TF8,007 TFL07FCK03TF8,007 SFCJ0LFCJ0TF8,007 SF8I02MF8I07SF8,007 SFI01BFFEC07EJ03SF8,007 RFCJ0EF88001L0SF8,007 RF8I0JFJ01ECI07RF8,007 RFI01IFM0BI03RF8,007 QFEI07FF8L07FC001RF8,007 QFC001IFCM092I0RF8,007" \
                  "QF8003FF1O038007QF8,007 QFI06FFP03C007QF8,007 QFI0IFP01F003QF8,007 PFEI0IF8P0B001QF8,007 PFC005FFEQ0BC00QF8,007 PF8007FFEQ07C007PF8,007 PF80077FCQ05F007PF8,007 PFI0FFD8Q097003PF8,007 OFE003IFS0FC03PF8,007 OFE007IFS0FC01PF8,007 OFC00IFES0FE00PF8,007 OFC00IFT07F00PF8,007" \
                  "OF801IFT07F807OF8,007 OF803IFT03F807OF8,007 OF003IFT03FC07OF8,007 OF007EFFT03FC03OF8,007 NFE007IFT03FC03OF8,007 NFE007IFT03FE01OF8,007 NFE00IFET01FF01OF8,007 NFC01IFET01FF00OF8,007 NFC01IFCT03FF80OF8,: 007 NF801IFCU0FFC07NF8,007 NF803IFET01FFC07NF8,007" \
                  "NF803IFCT03FFC07NF8,007 NF803IFCT03FFE07NF8,007 NF007IFCT03FFE03NF8,: 007 NF00JFCT03FFE03NF8,007 NF00JFCT03IF03NF8,007 NF00JF8T07IF03NF8,007 MFE00JF8T07IF01NF8,007 MFE00JFET03IF01NF8,007 MFE00JFCT03IF01NF8,007 MFE00JFCT07IF81NF8,::007 MFE00JFCT0JF81NF8,007" \
                  "MFE00JFET0JF81NF8,007 MFE00KFS01JF81NF8,007 MFE01JFCS03JF81NF8,007 MFE01KF8R03JF81NF8,007 MFE00KFCR07JF81NF8,: 007 MFE00KFCR0KF81NF8,007 MFE00LFQ01KF81NF8,007 MFE00LFQ03KF01NF8,: 007 MFE00LFCP07KF01NF8,007 MFE00LFCP0LF01NF8,007 MFE00LFEO03LF01NF8,007 NF00MFO07LF03NF8,007" \
                  "NF007LF8N0LFE03NF8,007 NF007MFM03LFE03NF8,007 NF007MFM0MFE03NF8,007 NF003MFCK03MFE03NF8,007 NF803NFCI01NFC07NF8,007 NF803OF8F7OFC07NF8,007 NF801gGFC07NF8,007 NFC01gGF807NF8,007 NFC00gGF80OF8,007 NFC00gGF00OF8,007 NFE007gF00OF8,007" \
                  "NFE007YFE01OF8,: 007 OF003YFC01OF8,007 OF003YFC03OF8,007 OF801YF803OF8,007 OF800YF007OF8,: 007 OFC007WFE00PF8,007 OFC003WFC00PF8,007 OFE001WFC01PF8,007 PFI0WF801PF8,007 PFI07VF003PF8,007 PF8007UFE007PF8,007" \
                  "PFC003UFC007PF8,007 PFC001UF800QF8,007 PFEI0UF001QF8,007 QFI03SFE001QF8,007 QFI01SF8003QF8,007 QF8I0SFI07QF8,007 QFCI03QFEI0RF8,007 QFEI01QF8001RF8,007 RFJ07OFEI03RF8,007 RF8I01OF8I07RF8,007 RFCJ07MFEJ0SF8,007" \
                  "RFEK0MFJ01SF8,007 SF8J01KF8J03SF8,007 SFCK01IFL07SF8,007 SFET01TF8,007 TF8S03TF8,007 TFCS0UF8,007U FR03UF8,007U FCQ07UF8,007 VFP03VF8,007 VFEO0WF8,007 WFCM07WF8,007" \
                  "XF8K03XF8,007 YFCI07YF8,007 YFCI0gF8,::::::::::::::003 YFCI0gF,:::001 YFCI0YFE,: 001" \
                  "YFCI0YFC,I0YFCI0YFC,I07XFCI0YF8,: I03XFCI0YF,I01XFCI0XFE,J0XFCI0XFC,J07WFCI0XF8,J03WFCI0XF,J01WFCI0WFE,K07VFCI0WF8,K01VFCI0VFE,L07UFCI0VF8,,:::::::: ^ FS" \
                  "^FO260,50^FDECO S.A.^FS" \
                  "^CF0,30" \
                  "^FO260,115^FDValparaiso, Chile^FS" \
                  "^FO260,155^FDAv. Central N490^FS" \
                  "^FO260,195^FD 32 2293300 - eco@ecosa.cl^FS" \
                  "^FO50,250^GB700,3,3^FS" \
                  "^FX SECCION DATOS CLIENTE." \
                  "^CF0,30" \
                  "^FO50,300^FDNombre: " + razon_social + "^FS" \
                  "^FO50,340^FDDireccion: " + direccion_1 + "^FS" \
                  "^FO50,380^FD" + direccion_2 + "^FS" \
                  "^FO50,420^FDComuna: " + comuna + "^FS" \
                  "^FO50,460^FDCiudad: " + ciudad + "^FS" \
                  "^FO50,500^GB700,3,3^FS" \
                  "^FX DATOS DEL TRANSPORTE." \
                  "^CF0,30" \
                  "^FO50,540^FDPedido: " + codigo_nota_de_venta + "^FS" \
                  "^FO50,580^FDTransporte: " + transporte_2 + "^FS" \
                  "^FO50,620^FDBultos: " + cantidad_bultos + "^FS" \
                  "^FO50,660^FDFecha Despacho: " + str(fecha_despacho)[:19] + "^FS" \
                  "^FX SECCION CODIGO DE BARRAS." \
                  "^BY5,2,60" \
                  "^FO100,693^BY2^BC,100,Y,N,N,A^FD"+"ORDEN-ENTREGA/"+codigo_nota_de_venta+"^FS" \
                  "^XZ"



        nombre_impresora = "Zebra_Technologies_ZTC_GK420t"
        # nombre_impresora = "FS-1320D"
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(host, port, username, password)
        #ssh.exec_command(command)
        #ssh.exec_command('lp -d ' + nombre_impresora + ' -o raw /home/' + nombre_archivo)
        #ssh.exec_command('rm -rf /home/' + nombre_archivo)

        # Seccion impresion documento
        ssh.exec_command('touch /home/'+nombre_archivo)
        ssh.exec_command("echo " + command + " > /home/" + nombre_archivo)
        ssh.exec_command('lp -d ' + nombre_impresora + ' -o raw /home/' + nombre_archivo)
        #ssh.exec_command('rm -rf /home/' + nombre_archivo)
        #shh.close()

    return "IMPRESION ETIQUETA ORDEN DE DESPACHO EJECUTADA CON EXITO!"


@app.route('/impresion_picking/<nota_de_venta_api>', methods=['GET', 'POST'])
def impresion_picking(nota_de_venta_api):
    nota_de_venta_api = nota_de_venta_api.replace("_", "/")

    # SECCION IMPRESION

    host = "192.168.0.187"
    port = 22
    username = "root"
    password = "mcpt14a12"
    nombre_archivo = str(uuid.uuid4()) + ".zpl"
    command = "^XA" \
              "^FX" \
              "^CF0,35" \
              "^FO120,130^FD" + nota_de_venta_api + "^FS" \
              "^FX" \
              "^FO30,250" \
              "^BY2" \
              "^BC,150,Y,N,N,A^FD" + nota_de_venta_api + \
              "^FS" \
              "^XZ"

    nombre_impresora = "Zebra_Technologies_ZTC_GK420t"
    # nombre_impresora = "FS-1320D"
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(host, port, username, password)
    #ssh.exec_command(command)
    #ssh.exec_command('lp -d ' + nombre_impresora + ' -o raw /home/' + nombre_archivo)
    #ssh.exec_command('rm -rf /home/' + nombre_archivo)
    
    # Seccion impresion documento
    ssh.exec_command('touch /home/'+ nombre_archivo)
    ssh.exec_command("echo " + command + " > /home/"+ nombre_archivo)
    #ssh.exec_command('lp -d ' + nombre_impresora + ' -o raw /home/' + nombre_archivo)
    #ssh.exec_command('rm -rf /home/' + nombre_archivo)
    #ssh.close()

    return "IMPRESION DE ETIQUETA PICKING CON EXITO! " + nota_de_venta_api


@app.route('/impresion_picking_final/<nota_de_venta_api>', methods=['GET', 'POST'])
def impresion_picking_final(nota_de_venta_api):
    nota_de_venta_api = nota_de_venta_api.replace("_", "/")
    nota_de_venta_api_2 = nota_de_venta_api.replace("NOTA-VENTA/", "")

    # SECCION IMPRESION

    host = "192.168.0.187"
    port = 22
    username = "root"
    password = "mcpt14a12"
    nombre_archivo = str(uuid.uuid4()) + ".zpl"
    command = "^XA" \
              "^FX" \
              "^CF0,35" \
              "^FO120,130^FD" + nota_de_venta_api + "^FS" \
              "^FX" \
              "^FO150,250" \
              "^BY2" \
              "^BC,150,Y,N,N,A^FD" + nota_de_venta_api_2 + \
              "^FS" \
              "^XZ"

    nombre_impresora = "Zebra_Technologies_ZTC_GK420t"
    # nombre_impresora = "FS-1320D"
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(host, port, username, password)
    #ssh.exec_command(command)

    # Seccion impresion documento
    ssh.exec_command('touch /home/'+ nombre_archivo)
    ssh.exec_command("echo " + command + " > /home/"+ nombre_archivo)
    ssh.exec_command('lp -d ' + nombre_impresora + ' -o raw /home/' + nombre_archivo)
    #ssh.exec_command('rm -rf /home/' + nombre_archivo)
    ssh.close()

    return "IMPRESION DE ETIQUETA PICKING FINAL CON EXITO! " + nota_de_venta_api



@app.route('/impresion_bultos/<nota_de_venta_api>', methods=['GET', 'POST'])
def impresion_bultos(nota_de_venta_api):

    nota_de_venta_api = "%["+nota_de_venta_api.replace('-','')+"]%"
    conexion = psycopg2.connect("host=192.168.0.154 dbname=InventorySystem user=odoo14 password=  ")
    cursor = conexion.cursor()
    query = """SELECT name FROM stock_quant_package WHERE name LIKE  %s """
    cursor.execute(query, (nota_de_venta_api,))
    datos = cursor.fetchall()



    for value in datos:
        print(value[0])
        host = "192.168.0.205"
        port = 22
        username = "root"
        password = "mcpt14a12"
        nombre_impresora = "Zebra_Technologies_ZTC_GK420t"
        # nombre_impresora = "FS-1320D"
        nombre_archivo = str(uuid.uuid4()) + ".zpl"
        command = "^XA^FX^CF0,35^FO250,130^FD" + value[0] + "^FS^FX^FO50,250^BY3,5^BC,150,Y,N,N,A^FD" + value[0] + "^FS^XZ"

        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(host, port, username, password)

        # Seccion impresion documento
        ssh.exec_command('touch /home/' + nombre_archivo)
        ssh.exec_command("echo " + command + " > /home/" + nombre_archivo)
        ssh.exec_command('lp -d ' + nombre_impresora + ' -o raw /home/' + nombre_archivo)

    return "IMPRESION DE BULTOS CON EXITO! " + nota_de_venta_api




# ESTA API NO ESTA FUNCIONANDO DEL TODO BIEN, HAY QUE SOLUCIONARLO
@app.route('/delete_packages/<name_package>', methods=['GET', 'POST'])
def eliminacion_paquetes(name_package):
    name_package = name_package.replace("PACK_", "PACK/")
    lista_paquetes = name_package.split(",")

    def listToString(s):
        str1 = " "
        return (str1.join(s))

    print(lista_paquetes)
    con = psycopg2.connect("host=192.168.0.154 dbname=InventorySystem user=postgres password=  ")
    cur = con.cursor()

    for key, value in enumerate(lista_paquetes):

        print(value)

        cur.execute("SELECT id FROM stock_quant_package WHERE name = %s;", (value,))
        datos = cur.fetchall()
        print(datos)

        if bool(datos) == False:
            cur.execute("DELETE FROM stock_quant_package WHERE name = %s", (value,))
            rows_deleted = cur.rowcount
            con.commit()
        else:
            id_package = ''.join(str(datos[0])).replace(",)", "").replace("(", "").replace(")", "")

            nombre_pack = value
            posicion_inicial = nombre_pack.find("[") + 1
            posicion_final = nombre_pack.find("]")
            nombre_pack_sub = nombre_pack[posicion_inicial:posicion_final]
            nombre_pack_len = len(nombre_pack_sub)
            ultimo_numero = nombre_pack_sub[nombre_pack_len - 1:]
            codigo_nota_venta = "PICKING-FINAL/NOTA-VENTA/[" + nombre_pack_sub[:-1] + "-" + ultimo_numero + "]"

            nota = ""
            nada = None

            cur.execute("DELETE FROM stock_package_level WHERE package_id = %s", (id_package,))
            cur.execute("UPDATE stock_move_line SET result_package_id = %s WHERE result_package_id = %s",
                        (nada, id_package,))
            cur.execute("DELETE FROM stock_quant_package WHERE name = %s", (value,))
            cur.execute("UPDATE stock_picking SET note = %s WHERE name = %s", (nota, codigo_nota_venta))
            con.commit()
            rows_deleted = cur.rowcount
            con.commit()

    return "ELIMINACION DE PAQUETES CON EXITO!"






@app.route('/insert_packages/<code_package>', methods=['GET', 'POST'])
def insert_paquetes(code_package):
    codigo_nota_venta = code_package

    con = psycopg2.connect("host=192.168.0.200 dbname=imatronix_ecosa user=postgres password=  ")
    cur = con.cursor()
    cur.execute("SELECT _id,transporte_2 FROM nota_de_venta WHERE codigo = %s", (code_package,))
    datos = cur.fetchall()

    id_nota_de_venta = ""
    transporte_2 = ""
    for _id, transporte_2 in datos:
        id_nota_de_venta = _id
        transporte_2 = transporte_2

    con_154 = psycopg2.connect("host=192.168.0.154 dbname=InventorySystem user=postgres password=  ")
    cur_154 = con_154.cursor()
    name = "PICKING-FINAL/NOTA-VENTA/[" + code_package + "]"
    cur_154.execute("SELECT note FROM stock_picking WHERE name = %s", (name,))
    result = cur_154.fetchall()

    nota_de_venta_id = str(datos).strip('[]').replace("(", "").replace(")", "").replace(",", "")
    nota = str(result).strip('[]').replace("(", "").replace(")", "").replace(",", "")

    # INSERT CRECION DE BULTOS
    embalador = 34
    fecha = str(datetime.now())[:-7]
    transporte_1 = None
    transporte_2 = transporte_2
    peso_total_bruto_kg = 0
    peso_total_neto_kg = 0
    peso_total_nota_de_venta_kg = 0
    bodega = 1
    cantidad_de_bultos = 0
    nota_de_venta_id = id
    transporte_3 = "."


    #INICIO INSEERT
    cur.execute("INSERT INTO creacion_de_bultos (embalador, fecha, transporte_1, transporte_2, peso_total_bruto_kg, peso_total_neto_kg, peso_total_nota_de_venta_kg, bodega, cantidad_de_bultos, nota_de_venta_id, transporte_3)"
        "VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
        (embalador, fecha, transporte_1, transporte_2, peso_total_bruto_kg, peso_total_neto_kg,peso_total_nota_de_venta_kg, bodega, cantidad_de_bultos, nota_de_venta_id, transporte_3))
    con.commit()
    # FIN INSERT CREACION DE BULTOS


    # INSERT BULTOS
    cur.execute("SELECT _id FROM creacion_de_bultos WHERE nota_de_venta_id = %s", (id_nota_de_venta,))
    res = cur.fetchall()
    id_creacion_bulto = str(res).replace(",)]", "").replace("[(", "")

    suma_pesajes = []
    for i in nota.split("\\n"):
        if (len(i) > 4):
            paquete_bruto = i.replace("'", "")
            valor_eliminacion = str("PACK/[" + str(codigo_nota_venta).replace("-", "") + "]-")
            valores_bruto = paquete_bruto.replace(valor_eliminacion, "").replace(" → PESAJE ", "").replace("KG", "")

            peso_bruto_kg = float(valores_bruto[2:])
            peso_neto_kg = str(valores_bruto[2:])
            creacion_de_bultos_id = str(id_creacion_bulto)
            bulto_n = str(valores_bruto[:1])
            suma_pesajes.append(peso_bruto_kg)

            cur.execute("INSERT INTO bulto (bulto_n, peso_bruto_kg, peso_neto_kg, creacion_de_bultos_id) VALUES(%s,%s,%s,%s)",(bulto_n, peso_bruto_kg, peso_neto_kg, creacion_de_bultos_id))
            con.commit()
    # FIN INSERT BULTOS

    # UPDATE CREACION DE BULTOS
    pesaje_total = sum(suma_pesajes)
    peso_nota_de_venta = 0
    estado = "b"
    print(id_nota_de_venta)
    cur.execute("UPDATE creacion_de_bultos SET peso_total_bruto_kg = %s WHERE nota_de_venta_id = %s ",(pesaje_total, id_nota_de_venta))
    cur.execute("UPDATE creacion_de_bultos SET peso_total_neto_kg = %s WHERE nota_de_venta_id = %s ",(pesaje_total, id_nota_de_venta))
    cur.execute("UPDATE creacion_de_bultos SET peso_total_nota_de_venta_kg = %s WHERE nota_de_venta_id = %s ",(peso_nota_de_venta, id_nota_de_venta))
    cur.execute("UPDATE nota_de_venta SET estado = %s WHERE _id = %s", (estado, id_nota_de_venta))
    con.commit()
    # FIN UPDATE CREACION DE BULTOS

    return "ADICION DE PAQUETES CON EXITO!"








@app.route('/codigos_de_barra_truper/<codigo_origen_producto>', methods=['GET', 'POST'])
def codigos_de_barra_truper(codigo_origen_producto):
    conexion = psycopg2.connect("host=192.168.0.154 dbname=InventorySystem user=postgres password=  ")
    cursor = conexion.cursor()
    consulta = ''' SELECT product_template.id, product_template.name, codigos_truper.codigo, codigos_truper.codigo_de_barra, codigos_truper.dun_14, codigos_truper.piezas_dun14, codigos_truper.dun_16, codigos_truper.piezas_dun_16 FROM codigos_truper JOIN product_template ON product_template.description = codigos_truper.codigo WHERE codigos_truper.codigo = %s; '''
    cursor.execute(consulta, (codigo_origen_producto,))
    datos = cursor.fetchall()

    create_uid = 2
    write_uid = 2
    sequence = 1
    company_id = 1
    cantidad_piezas = ""
    name = ""
    fecha_actual = str(datetime.datetime.now())
    # FIN OBTENCIÓN DE DATOS

    # EN ESTA SECCION SE TRABAJA CON LA TABLA "product_packaging" → esto es importante para ver el futuro funcionamiento.
    for id, name, codigo, codigo_de_barra, dun_14, piezas_dun14, dun_16, piezas_dun_16 in datos:

        if (dun_14 == "" and dun_16 != ""):
            select_limitante_1 = ''' SELECT * FROM product_packaging WHERE barcode = %s '''
            argument = (dun_16,)
            cursor.execute(select_limitante_1, argument)
            result = cursor.fetchall()

            if bool(result) == False:
                print(
                    "No hay datos en la tabla product_packaging referentes al dun16 por ende se procede a ejecutar el insert ↓")
                cantidad_piezas = piezas_dun_16
                insert_packaging = f''' INSERT INTO product_packaging (name, sequence, product_id, qty, barcode, company_id, create_uid, create_date, write_uid, write_date) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s); '''
                values_insert = (
                    name + " DUN 16", sequence, id, cantidad_piezas, dun_16, company_id, create_uid, fecha_actual,
                    write_uid, fecha_actual)
                cursor.execute(insert_packaging, values_insert)
                conexion.commit()
                count = cursor.rowcount
                print(count)

            else:
                print(
                    "Hay datos en la tabla product_packaging referentes al dun16 por ende se procede no se ejecuta el insert6")






        elif (dun_16 == "" and dun_14 != ""):
            select_limitante_2 = ''' SELECT * FROM product_packaging WHERE barcode = %s '''
            argument = (dun_14,)
            cursor.execute(select_limitante_2, argument)
            result = cursor.fetchall()

            if bool(result) == False:
                print(
                    "No hay datos en la tabla product_packaging referentes al dun14 por ende se procede a ejecutar el insert ↓")
                cantidad_piezas = piezas_dun14
                insert_packaging = f''' INSERT INTO product_packaging (name, sequence, product_id, qty, barcode, company_id, create_uid, create_date, write_uid, write_date) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s); '''
                values_insert = (
                    name + " DUN 14", sequence, id, cantidad_piezas, dun_14, company_id, create_uid, fecha_actual,
                    write_uid, fecha_actual)
                cursor.execute(insert_packaging, values_insert)
                conexion.commit()
                count = cursor.rowcount
                print(count)
            else:
                print(
                    "Hay datos en la tabla product_packaging referentes al dun14 por ende se procede no se ejecuta el insert")







        elif (dun_16 != "" and dun_14 != ""):
            select_limitante_3 = ''' SELECT * FROM product_packaging WHERE barcode = %s '''
            argument = (dun_14,)
            cursor.execute(select_limitante_3, argument)
            result_1 = cursor.fetchall()

            select_limitante_4 = ''' SELECT * FROM product_packaging WHERE barcode = %s '''
            argument_2 = (dun_16,)
            cursor.execute(select_limitante_4, argument_2)
            result_2 = cursor.fetchall()

            if bool(result_1) == False and bool(result_2) == False:

                cantidad_piezas = piezas_dun14
                insert_packaging = f''' INSERT INTO product_packaging (name, sequence, product_id, qty, barcode, company_id, create_uid, create_date, write_uid, write_date) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s); '''
                values_insert = (
                    name + " DUN 14", sequence, id, cantidad_piezas, dun_14, company_id, create_uid, fecha_actual,
                    write_uid, fecha_actual)
                cursor.execute(insert_packaging, values_insert)
                conexion.commit()
                count = cursor.rowcount
                print(count)

                cantidad_piezas_2 = piezas_dun_16
                insert_packaging = f''' INSERT INTO product_packaging (name, sequence, product_id, qty, barcode, company_id, create_uid, create_date, write_uid, write_date) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s); '''
                values_insert = (
                    name + " DUN 16", sequence, id, cantidad_piezas_2, dun_16, company_id, create_uid, fecha_actual,
                    write_uid, fecha_actual)
                cursor.execute(insert_packaging, values_insert)
                conexion.commit()
                count = cursor.rowcount
                print(count)

            else:
                print("Hay datos tanto de dun 14 como de dun 16 por ende no se ejecuta el insert")

    ####################################################################################################################################################################################################################################

    # En esta seccion se trabaja con la tabla "product_barcode_multi"
    for id, name, codigo, codigo_de_barra, dun_14, piezas_dun14, dun_16, piezas_dun_16 in datos:

        if (dun_14 == ""):
            select_limitante_2 = ''' SELECT * FROM product_barcode_multi WHERE name = %s '''
            argument = (dun_16,)
            cursor.execute(select_limitante_2, argument)
            result = cursor.fetchall()

            if bool(result) == False:
                insert = f'''INSERT INTO product_barcode_multi(name, product_id, create_uid, create_date, write_uid, write_date) VALUES(%s,%s,%s,%s,%s,%s);'''
                datos_insert = (dun_16, id, create_uid, fecha_actual, write_uid, fecha_actual)
                cursor.execute(insert, datos_insert)
                conexion.commit()
                count = cursor.rowcount
                print(count, "DUN 16 product_barcode_multi AGREGADO CON EXITO ")




        elif (dun_16 == ""):
            select_limitante_2 = ''' SELECT * FROM product_barcode_multi WHERE name = %s '''
            argument = (dun_14,)
            cursor.execute(select_limitante_2, argument)
            result = cursor.fetchall()

            if bool(result) == False:
                insert = f'''INSERT INTO product_barcode_multi(name, product_id, create_uid, create_date, write_uid, write_date) VALUES(%s,%s,%s,%s,%s,%s);'''
                datos_insert = (dun_14, id, create_uid, fecha_actual, write_uid, fecha_actual)
                cursor.execute(insert, datos_insert)
                conexion.commit()
                count = cursor.rowcount
                print(count, "DUN 14 product_barcode_multi AGREGADO CON EXITO ")




        elif (dun_14 != "" and dun_16 != ""):

            select_limitante_3 = ''' SELECT * FROM product_barcode_multi WHERE name = %s '''
            argument = (dun_14,)
            cursor.execute(select_limitante_3, argument)
            result_1 = cursor.fetchall()

            select_limitante_4 = ''' SELECT * FROM product_barcode_multi WHERE name = %s '''
            argument_2 = (dun_16,)
            cursor.execute(select_limitante_4, argument_2)
            result_2 = cursor.fetchall()

            if bool(result_1) == False and bool(result_2) == False:
                insert = f'''INSERT INTO product_barcode_multi(name, product_id, create_uid, create_date, write_uid, write_date) VALUES(%s,%s,%s,%s,%s,%s);'''
                datos_insert = (dun_14, id, create_uid, fecha_actual, write_uid, fecha_actual)
                cursor.execute(insert, datos_insert)
                conexion.commit()
                count = cursor.rowcount
                # print(count, "MOVIMIENTOS ACTUALIZADOS ACTUALIZADOS CON EXITO")

                insert_2 = f'''INSERT INTO product_barcode_multi(name, product_id, create_uid, create_date, write_uid, write_date) VALUES(%s,%s,%s,%s,%s,%s);'''
                datos_insert = (dun_16, id, create_uid, fecha_actual, write_uid, fecha_actual)
                cursor.execute(insert_2, datos_insert)
                conexion.commit()
                count = cursor.rowcount

    return "API CODIGO DE BARRAS TRUPER STATUS 200!"


# Gestion del documento XLSX.
# Es necesario pasar como parametro el nombre del archivo en el directorio ubicaciones.
# Esta API lo que hace es generar
@app.route('/crear_ubicaciones/<nombre_tabla>', methods=['GET', 'POST'])
def crear_ubicaciones(nombre_tabla):
    tabla = str(nombre_tabla).upper() + ".xlsx"
    wb = load_workbook(filename='/home/ubicaciones/' + tabla)
    nombre_sheet = wb.sheetnames[0]
    ws = wb[nombre_sheet]

    data_rows = []
    for iteration, row in enumerate(ws['A2':'F1300']):

        data_cols = []
        for cell in row:
            data_cols.append(cell.value)
        data_rows.append(data_cols)

    # Se convierte la hoja de calculo a un DataFrame con la libreria 'Pandas'
    df = pd.DataFrame(data_rows)

    # conexion a la base de datos
    con_154 = psycopg2.connect("host=192.168.0.154 dbname=InventorySystem user=postgres password=  ")
    cur_154 = con_154.cursor()

    # Eliminacion de la tabla si es que existe.
    drop_table_locations = f''' DROP TABLE IF EXISTS {str(nombre_tabla)}; '''
    cur_154.execute(drop_table_locations)

    # creacion de la tabla.
    craeate_table_locations = f''' CREATE TABLE {str(nombre_tabla)}(
                                 SECCION VARCHAR(100),
                                 PASILLO VARCHAR(100),
                                 FILA VARCHAR(100),
                                 COLUMNA VARCHAR(100),
                                 NIVEL  VARCHAR(100),
                                 PRODUCTO VARCHAR(100)
                                 ); '''
    cur_154.execute(craeate_table_locations)
    con_154.commit()

    for index, row in df.iterrows():
        if (row[0] != '' and row[0] != None and row[0] != ' '):
            # row = columna de la hoja de calculo
            seccion = row[0]
            pasillo = row[1]
            fila = row[2]
            columna = row[3]
            nivel = row[4]
            producto = row[5]

            insert_table_locations = f''' INSERT INTO {str(nombre_tabla)} (seccion, pasillo, fila, columna, nivel, producto) VALUES(%s,%s,%s,%s,%s,%s) '''
            data_insert = (seccion, pasillo, fila, columna, nivel, producto)
            cur_154.execute(insert_table_locations, data_insert)
            con_154.commit()

    # Eliminacion de datos basura
    delete_null = f''' DELETE FROM {str(nombre_tabla)} WHERE producto IS NULL OR producto = '' '''
    cur_154.execute(delete_null)
    con_154.commit()

    return "Gestión ubicaciones ejecutada con exito"


# Esta API actualiza ta tabla "description_pickingout" de la tabla "product_template", con el objetivo de ordenas las ordenes de picking con la API "ajax_200"
@app.route('/actualizar_ubicaciones/<nombre_tabla>', methods=['GET', 'POST'])
def actualizar_ubicaciones(nombre_tabla):
    # Obtención de datos del pasillo correspondiente
    conexion = psycopg2.connect(user='odoo14', password='', host='192.168.0.154', database='odoo14')
    cursor = conexion.cursor()
    query = f''' SELECT * FROM {nombre_tabla} '''
    cursor.execute(query)
    datos = cursor.fetchall()

    # Parametros a llenar
    seccion = ""
    pasillo = ""
    fila = ""
    columna = ""
    nivel = ""
    producto = ""

    for index, tupla in enumerate(datos):
        tupla_array = tupla

        for index2, v in enumerate(tupla_array):
            if (index2 == 0):
                seccion = v
            if (index2 == 1):
                pasillo = v
            if (index2 == 2):
                fila = v
            if (index2 == 3):
                columna = v
            if (index2 == 4):
                nivel = v
            if (index2 == 5):
                producto = v

        ubicacion = "SECCION=" + seccion + " PASILLO=" + pasillo + " FILA=" + fila + " COLUMNA=" + columna + " NIVEL=" + nivel
        conn = psycopg2.connect(database="odoo14", user="postgres", password='', host='192.168.0.154', port='5432')
        cursor = conn.cursor()
        cursor.execute("UPDATE product_template SET description_pickingout=(%s) WHERE SUBSTRING(name,0,6)=(%s)",
                       (ubicacion, str(producto),));
        cursor.execute("UPDATE product_template SET description_pickingout=(%s) WHERE SUBSTRING(name,0,7)=(%s)",
                       (ubicacion, str(producto),));
        conn.commit()
        # cursor.close()

    return "actualizacion description_pickingout ejecutada con exito"


@app.route('/productos_sin_codigo_de_barra/<hoja_de_calculo>', methods=['GET', 'POST'])
def productos_sin_codigo_de_barra(hoja_de_calculo):
    hoja_de_calculo_final = str(hoja_de_calculo) + '.xlsx'
    # Gestion del documento XLSX -> ES NECESARIO QUE SEA EN ESTE FORMATO, SI NO LO ES NO VA A FUNCIONAR!.
    # Como parametro de podría pasar la ubicacion del archivo 'XLSX'
    wb = load_workbook(filename='/home/CatalogosTruper/' + hoja_de_calculo_final)
    nombre_sheet = wb.sheetnames[0]
    ws = wb[nombre_sheet]

    data_rows = []
    for iteration, row in enumerate(ws['C4':'AD18000']):
        data_cols = []
        for cell in row:
            data_cols.append(cell.value)
        data_rows.append(data_cols)

    # Se convierte con la libreria 'Pandas' a un DataFrame
    df = pd.DataFrame(data_rows)
    # Fin de la gestion 'XLSX'

    dict = dict()

    # Recopilacion de datos a trabajar
    codigo_prod_truper = ""
    index_codigo_prod_truper = ""
    clave = ""
    index_clave = ""
    descripcion = ""
    index_descripcion = ""
    numero_familia = ""
    index_numero_familia = ""
    marca = ""
    index_marca = ""
    dun_14 = ""
    index_dun_14 = ""
    piezas_dun14 = ""
    index_piezas_dun14 = ""
    dun_16 = ""
    index_dun_16 = ""
    piezas_dun_16 = ""
    index_piezas_dun_16 = ""
    peso_por_pieza = ""
    index_peso_por_pieza = ""
    volumen_cm3 = ""
    index_volumen_cm3 = ""
    codigo_de_barras = ""
    index_codigo_de_barras = ""
    # Fin recopilación de datos

    # Selección de columnas con las cuales vamos a trabajar.(solo las que necesitamos).
    for index, row in df.iterrows():

        # GENERACIÓN DE UNA LISTA DE LAS COLUMNAS Y SU POSICIÓN RESPECTIVA
        a_list = str(row).split("  ")
        filtro = filter(lambda x: x != "", a_list)
        lista_sin_void = list(filtro)
        new_lst = [x[:-1] for x in lista_sin_void]
        rep = []
        for x in lista_sin_void:
            rep.append(x.replace("\n", "  "))
        lista_final = list(rep)
        del lista_final[0]
        len_lista = int(len(lista_final)) - 1
        # del lista_final[len_lista]
        # print(lista_final)
        # FIN GENERACIÓN DE COLUMNAS

        # CREACION DE UN ARRAY ASOCIATIVO CON EL NOMBRE DE LA COLUMNA Y SU POSICION
        if (index == 0):
            for index, datos in enumerate(lista_final):
                longitud_nombre_columna = len(datos)
                sub = int(longitud_nombre_columna) - 2
                nombre_final = str(datos[0:sub]).lower().replace('ó', 'o')
                posicion_final = datos[-2:]
                if (posicion_final.isnumeric() == True):
                    posicion_final = str(int(datos[-2:]) - 1)
                else:
                    posicion_final = int(index)

                dict[nombre_final.replace('  name: 0, dtype: obje', '')] = posicion_final

            # print(dict)
            nombre_columna = ""
            for key, value in dict.items():

                nombre_columna = key
                nombre_columna_lstrip = nombre_columna.lstrip()
                nombre_columna_rstrip = nombre_columna_lstrip.rstrip()
                nombre_columna_final = nombre_columna_rstrip.replace('  name: 0, dtype: obje', '')

                # ASIGNACION DE VALORES PARA EJECUTAR EL CÓDIGO
                if (nombre_columna_final.find('codigo') >= 0):
                    if (nombre_columna_final == 'codigo'):
                        codigo_prod_truper = row[int(value)]
                        index_codigo_prod_truper = int(value)

                if (nombre_columna_final.find('clave') >= 0):
                    if (nombre_columna_final == 'clave'):
                        clave = row[int(value)]
                        index_clave = int(value)

                if (nombre_columna_final.find('descripcion') >= 0):
                    if (nombre_columna_final == 'descripcion'):
                        descripcion = row[int(value)]
                        index_descripcion = int(value)

                if (nombre_columna_final.find('familia') >= 0):
                    if (nombre_columna_final == 'no. familia'):
                        numero_familia = row[int(value)]
                        index_numero_familia = int(value)

                if (nombre_columna_final.find('marca') >= 0):
                    if (nombre_columna_final == 'marca'):
                        marca = row[int(value)]
                        index_marca = int(value)

                if (nombre_columna_final.find('dun 14') >= 0):
                    if (nombre_columna_final == 'dun 14'):
                        dun_14 = row[int(value)]
                        index_dun_14 = int(value)

                if (nombre_columna_final.find('caja') >= 0):
                    if (nombre_columna_final == 'caja' or nombre_columna_final == 'caja(inner)zem'):
                        piezas_dun14 = row[int(value)]
                        index_piezas_dun14 = int(value)

                if (nombre_columna_final.find('dun 16') >= 0):
                    if (nombre_columna_final == 'dun 16'):
                        dun_16 = row[int(value)]
                        index_dun_16 = int(value)

                if (nombre_columna_final.find('master') >= 0):
                    if (nombre_columna_final == 'master' or nombre_columna_final == 'master zcm'):
                        piezas_dun_16 = row[int(value)]
                        index_piezas_dun_16 = int(value)

                if (nombre_columna_final.find('peso x pza') >= 0):
                    if (nombre_columna_final == 'peso x pza'):
                        peso_por_pieza = row[int(value)]
                        index_peso_por_pieza = int(value)

                if (nombre_columna_final.find('volumen cm3') >= 0):
                    if (nombre_columna_final == 'volumen cm3'):
                        volumen_cm3 = row[int(value)]
                        index_volumen_cm3 = int(value)

                if (nombre_columna_final.find('codigo de barras') >= 0):
                    if (nombre_columna_final == 'codigo de barras'):
                        codigo_de_barras = row[int(value)]
                        index_codigo_de_barras = int(value)

    for index, row in df.iterrows():

        codigo_prod_truper = row[index_codigo_prod_truper]
        clave = row[index_clave]
        descripcion = row[index_descripcion]
        numero_familia = row[index_numero_familia]
        marca = row[index_marca]
        dun_14 = row[index_dun_14]
        piezas_dun14 = row[index_piezas_dun14]
        dun_16 = row[index_dun_16]
        piezas_dun_16 = row[index_piezas_dun_16]
        peso_por_pieza = row[index_peso_por_pieza]
        volumen_cm3 = row[index_volumen_cm3]
        codigo_de_barras = row[index_codigo_de_barras]
        column_11 = ""
        column_12 = ""
        column_13 = ""

        if (codigo_prod_truper != None):

            # Conexiones a las bases de datos.
            con_200 = psycopg2.connect("host=192.168.0.200 dbname=imatronix_ecosa user=postgres password=  ")
            cur_200 = con_200.cursor()
            query_200 = f''' SELECT codigo_prod_truper FROM codigos_de_barra_truper WHERE codigo_prod_truper = '{codigo_prod_truper}'  '''
            cur_200.execute(query_200)
            condicional_datos_server_200 = cur_200.fetchall()

            con_154 = psycopg2.connect("host=192.168.0.154 dbname=InventorySystem user=postgres password=  ")
            cur_154 = con_154.cursor()
            query_154 = f''' SELECT codigo FROM codigos_truper WHERE codigo = '{codigo_prod_truper}'  '''
            cur_154.execute(query_154)
            condicional_datos_server_154 = cur_154.fetchall()

            if bool(condicional_datos_server_200) == False or bool(condicional_datos_server_154) == False:
                print("SERVIDOR 200 -> No existe el valor en la base de datos -> " + str(codigo_prod_truper))
                cur_200.execute(
                    "INSERT INTO codigos_de_barra_truper (codigo_prod_truper, clave, codigo_de_barras, numero_familia, marca, dun_14, piezas_dun14, dun_16, piezas_dun_16, peso_por_pieza, volumen_cm3, column_11,column_12,column_13) VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                    (codigo_prod_truper, clave, codigo_de_barras, numero_familia, marca, dun_14, piezas_dun14, dun_16,
                     piezas_dun_16, peso_por_pieza, volumen_cm3, column_11, column_12, column_13,))
                con_200.commit()

                print("SERVIDOR 154 -> No existe el valor en la base de datos" + str(codigo_prod_truper))
                cur_154.execute(
                    "INSERT INTO codigos_truper (codigo, clave, codigo_de_barra, numero_familia, marca, dun_14, piezas_dun14, dun_16, piezas_dun_16, peso_por_pieza, volumen_cm3, column_11,column_12,column_13) VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                    (codigo_prod_truper, clave, codigo_de_barras, numero_familia, marca, dun_14, piezas_dun14, dun_16,
                     piezas_dun_16, peso_por_pieza, volumen_cm3, column_11, column_12, column_13,))
                con_154.commit()


            elif bool(condicional_datos_server_200) == True or bool(condicional_datos_server_154) == True:
                print("SERVIDOR 200 -> Existen datos referentes al siguiente valor " + str(codigo_prod_truper))
                # print(str(codigo_prod_truper))
                # continue

                print("SERVIDOR 154 -> Existen datos referentes al siguiente valor " + str(codigo_prod_truper))
                # print(str(codigo_prod_truper))
                # continue

    return "El producto X en su campo Y fue actualizado"


@app.route('/modificacion_nota_de_venta/<json>/<data>', methods=['GET', 'POST'])
def modificacion_nota_de_venta(json, data):
    def listToString(s):
        str1 = ""
        for ele in s:
            str1 += ele
        return str1

    # nota_de_venta = sys.argv[1]
    # json = sys.argv[2]

    nota_de_venta = data
    json = json

    nombre_orden_entrega = "<b>" + "BEC1/ORDEN-ENTREGA/" + nota_de_venta + "</b>"
    ##json_replace = json.replace('producto','"producto"').replace('cantidad','"cantidad"').replace('modificada','"modificada"').replace('diferencia','"diferencia"')
    ##json_replace = json.replace("'",'"')
    array = loads(json)

    ##CREACION  DE TABLA
    Tr_lista = []
    for value in array:
        nombre_producto = str(value["producto"]).replace("\u200B", " ").replace("\\u200B", " ").replace('\\u2215', '/')
        cantidad_original = str(value["cantidad"])
        cantidad_modificada = str(value["modificada"])
        diferencia = str(value["diferencia"])
        Tr = "<tr>" \
                  "<td style='border: 1px solid #ddd; padding: 8px;'>" + nombre_producto + "</td>" \
                  "<td style='border: 1px solid #ddd; padding: 8px;'>" + cantidad_original + "</td>" \
                  "<td style='border: 1px solid #ddd; padding: 8px;'>" + cantidad_modificada + "</td>" \
                  "<td style='border: 1px solid #ddd; padding: 8px;'>" + diferencia + "</td>" \
             "</tr>"
        Tr_lista.append(Tr)

    Tabla = "<table style='border-collapse: collapse; border: 1px solid #ddd; ont-family: Arial, Helvetica, sans-serif;border-collapse: collapse;width: 100%;'>" \
            "<tr>" \
                "<th style='border: 1px solid #ddd; padding: 8px; background-color: #4CAF50;'>" + "Producto" + "</th>" \
                "<th style='border: 1px solid #ddd; padding: 8px; background-color: #4CAF50;'>" + "Cantidad Original" + "</th>" \
                "<th style='border: 1px solid #ddd; padding: 8px; background-color: #4CAF50;'>" + "Cantidad Modificada" + "</th>" \
                "<th style='border: 1px solid #ddd; padding: 8px; background-color: #4CAF50;'>" + "Diferencia" + "</th>" \
            "</tr>" \
            + listToString(Tr_lista) + \
            "</table>"
    ##FIN CREACION TABLA.

    ##ENVIO DE MENSAJE
    receptores = ['fcumplido@ecosa.cl', 'franco.cumplido@gmail.com']
    # receptores = ['fcumplido@ecosa.cl', 'franco.cumplido@gmail.com','asessler@ecosa.cl','jbarraza@ecosa.cl','c.ahumada@ecosa.cl']
    for receptor in receptores:
        server = email_to.EmailServer('mail.ecosa.cl', 587, 'fcumplido@ecosa.cl', 'v2obzatm')
        message = server.message()
        message.add('# Nota de venta N° ' + nota_de_venta + ' fue modificada')
        message.add(
            '- Se modificaron automaticamente los valores de la orden ' + nombre_orden_entrega + ', porfavor agregar o quitar productos del pedido ' + nota_de_venta + ' se adjunta detalle en la siguiente tabla:' + Tabla)
        message.style = 'h1 { color: red}'
        message.send(receptor, 'ATENCIÓN!, nota de venta ' + nota_de_venta + ' modificada')
    ##FIN ENVIO DE MENSAJE

    return "Modificacion de la nota de veta ejecutada con exito"


if __name__ == '__main__':
    app.run(host='192.168.0.154', port=105)
